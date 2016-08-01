# 라이브러리 import
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment
import datetime

# 타이틀
print("인원 셔플과 출석부 제작을 해주는 프로그램입니다.\n")

people = [  "강려령", "강보경", "고요한", "김윤성",
            "김준호", "나정연", "송창현", "안우리",
            "유찬호", "최지수", "유태영", "이지윤",
            "이윤형", "이해석", "이찬하", "박정현",
            "윤형규", "민동준", "정예은", "문상호",
            "이혜인", "임다은", "임지민", "임형준",
            "최형규", "이호동", "홍창남", "김재훈", "함형우"]

print("현재 포함되어 있는 사람들입니다.")
print(people)

# 제외되는 사람 입력하고 셔플돌리기
print("\n띄어쓰기 단위로 끊어서 제외할 사람을 입력해주세요.(없으면 enter)")
out_of_group = input()

# 띄어쓰기 단위로 배열 나누기
out_of_group = out_of_group.split()

# 제외된 사람 기존 리스트에서 지우기
for i in out_of_group:
    people.remove(i)

print("다음 사람을 제외한 결과입니다.\n")
print(people)
print('\n')

# 인원 섞기
people = random.sample(people, len(people))

# 그룹별 출력
for i in range(len(people)):
    print(people[i]+ ', ', end="")
    if i != 0 and i % 4 == 3:
        print('\n')

# 출석부 export
print('\n해당 내용을 엑셀로 저장하시겠습니까?(y/n)')
answer = input()

if answer == 'y' or answer == "Y":
    wb = Workbook()

    ws = wb.active
    ws.column_dimensions["C"].width = 30.0

    ws['A1'] = "참여자"
    ws['A1'].alignment = Alignment(horizontal="center")
    ws['C3'] = datetime.datetime.now()

    # 칼럼별로 이름 writing
    for i in range(len(people)):
        index = 'A' + str(i+2)
        ws[index] = people[i]
        ws[index].alignment = Alignment(horizontal="center")

    # 날짜 형식 0801 처럼 만들기
    today = datetime.date.today()
    if today.month < 10:
        month = str('0') + str(today.month)
    if today.day < 10:
        day = str('0') + str(today.day)
    date = month + day

    wb.save('출석부'+ date +'.xlsx')
    print("저장되었습니다.")
